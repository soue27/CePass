import sqlite3
import openpyxl


wb = openpyxl.load_workbook('pass.xlsx', data_only=True)
ws = wb['sheet1']
stroka = []
spisok = []


def create_db(spisok1: list):
    # Создание таблицы базы данных, на основании полученного списка из парсинга эксель файла.
    # На входе имя базы данных и список загрузки
    connect = sqlite3.connect('sepass.sqlite3')
    cursor = connect.cursor()
    cursor.execute('CREATE TABLE IF NOT EXISTS pass (number, password)')
    sqlite_insert_query = """INSERT INTO pass
                                 (number, password)
                                 VALUES (?, ?)"""
    cursor.executemany(sqlite_insert_query, spisok1)
    connect.commit()
    connect.close()


for i in range(1, ws.max_row + 1):
    for col in range(1, 3):
        stroka.append(ws.cell(i, col).value)
    print(i)
    korteg = tuple(stroka)
    spisok.append(korteg)
    stroka = []


create_db(spisok)
print('all done')
