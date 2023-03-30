import sqlite3
import os
import sys
from loguru import logger
import openpyxl

global connect, cursor
# logger.add("base.log", format="{time} {level} {message}", level="DEBUG", rotation="1 week", compression="zip")


@logger.catch
async def open_db():
    global connect, cursor
    if os.path.exists('sepass.sqlite3'):
        connect = sqlite3.connect('sepass.sqlite3')
        cursor = connect.cursor()
        cursor.execute('''SELECT count(name) FROM sqlite_master WHERE type='table' AND name='pass' ''')
        if cursor.fetchone()[0] == 1:
            print('База данных подключена')
        else:
            print('что то пошло не так')
            # logger.error("нет таблицы в базе данных")
    else:
        print('что то пошло не так? обратитесь к разработчку, нет файла')
        # logger.error("нет файла с базой данных")


def close_db():
    connect.commit()
    connect.close()


# def create_db(spisok1: list):
#     # Создание таблицы базы данных, на основании полученного списка из парсинга эксель файла.
#     # На входе имя базы данных и список загрузки
#     connect = sqlite3.connect('sepass.sqlite3')
#     cursor = connect.cursor()
#     cursor.execute('CREATE TABLE IF NOT EXISTS pass (number, password)')
#     sqlite_insert_query = """INSERT INTO pass
#                                  (number, password)
#                                  VALUES (?, ?)"""
#     cursor.executemany(sqlite_insert_query, spisok1)
#     connect.commit()
#     connect.close()


def search_by_number(number: str):
    # Функция поиска по номеру прибора учета: str,
    results = cursor.execute(f"SELECT * FROM pass WHERE number LIKE '%{number.lower()}%'").fetchall()
    return results


def add_to_bd(newpass: list):
    # Добавляет новую пару №ПУ- Пароль в базу данных
    cursor.execute("INSERT INTO pass VALUES (?, ?);",
                   (newpass[0], newpass[1]))
    connect.commit()


def add_fromfile():
    error_count = 0
    try:
        wb = openpyxl.load_workbook("files\\forload.xlsx")
        ws = wb.active
    except:
        return 0, 0
    if ws.max_column - ws.min_column >= 1:
        for i in range(ws.min_row, ws.max_row + 1):
                if str(ws.cell(i, ws.min_column).value).isdigit() and str(ws.cell(i, ws.min_column + 1).value).isdigit():
                    cursor.execute("INSERT INTO pass VALUES (?, ?);", (ws.cell(i, ws.min_column).value, ws.cell(i, ws.min_column + 1).value))
                else:
                    error_count += 1
        connect.commit()
    else:
        return 0, 0
    return ws.max_row-ws.min_row + 1 - error_count, error_count


if __name__ == '__main__':
    print(add_fromfile())
